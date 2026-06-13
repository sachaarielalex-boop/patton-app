"""Broker Database page – CRE broker contacts and office lease contacts."""
import streamlit as st
import os


def _load_brokers():
    """Load broker data from Excel files."""
    try:
        import openpyxl
    except ImportError:
        st.warning("openpyxl not installed")
        return [], [], []

    brokers = []
    firms = []
    contacts = []

    def _find(name):
        """Find a data file in the app directory or cwd."""
        for d in [os.path.dirname(os.path.abspath(__file__)), os.getcwd()]:
            p = os.path.join(d, name)
            if os.path.exists(p):
                return p
        return None

    # Brokers list
    path1 = _find("brokers_data.xlsx")
    if not path1:
        st.warning("Broker data file not found.")
        return firms, brokers, contacts

    try:
        wb = openpyxl.load_workbook(path1, data_only=True)
        st.caption("Loaded: {} sheets {}".format(path1, wb.sheetnames))

        # South Florida CRE Firms
        if "South Florida CRE Firms" in wb.sheetnames:
            ws = wb["South Florida CRE Firms"]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if row[0]:
                    firms.append({
                        "firm": str(row[0] or "").strip(),
                        "category": str(row[1] or "").strip(),
                        "website": str(row[2] or "").strip(),
                        "status": str(row[3] or "").strip(),
                    })

        # Brokers List
        if "Brokers List" in wb.sheetnames:
            ws = wb["Brokers List"]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                name = str(row[2] or "").strip() if len(row) > 2 and row[2] else ""
                if not name:
                    continue
                brokers.append({
                    "name": name,
                    "email": str(row[3] or "").strip() if len(row) > 3 else "",
                    "mobile": str(row[4] or "").strip() if len(row) > 4 else "",
                    "type": str(row[6] or "").strip() if len(row) > 6 else "",
                    "organization": str(row[7] or "").strip() if len(row) > 7 else "",
                    "market": str(row[8] or "").strip() if len(row) > 8 else "",
                    "linkedin": str(row[9] or "").strip() if len(row) > 9 else "",
                })

        # Looking for Leases List
        if "Looking for Leases List" in wb.sheetnames:
            ws = wb["Looking for Leases List"]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                name = str(row[2] or "").strip() if len(row) > 2 and row[2] else ""
                if not name:
                    continue
                contacts.append({
                    "name": name,
                    "email": str(row[3] or "").strip() if len(row) > 3 else "",
                    "mobile": str(row[4] or "").strip() if len(row) > 4 else "",
                    "type": str(row[6] or "").strip() if len(row) > 6 else "",
                    "organization": str(row[7] or "").strip() if len(row) > 7 else "",
                    "market": str(row[8] or "").strip() if len(row) > 8 else "",
                    "linkedin": str(row[9] or "").strip() if len(row) > 9 else "",
                })
        wb.close()
    except Exception as e:
        st.error("Error loading brokers: {}".format(e))

    # Coral Gables Office Lease contacts
    path2 = _find("coral_gables_data.xlsx")
    if path2:
        try:
            wb2 = openpyxl.load_workbook(path2, data_only=True)
            ws2 = wb2[wb2.sheetnames[0]]
            for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, values_only=True):
                if row[0]:
                    contacts.append({
                        "name": str(row[0] or "").strip(),
                        "type": str(row[1] or "").strip() if len(row) > 1 else "",
                        "market": str(row[2] or "").strip() if len(row) > 2 else "",
                        "organization": str(row[0] or "").strip(),
                        "email": "",
                        "mobile": str(row[4] or "").strip() if len(row) > 4 else "",
                        "address": str(row[3] or "").strip() if len(row) > 3 else "",
                        "website": str(row[5] or "").strip() if len(row) > 5 else "",
                        "linkedin": "",
                    })
            wb2.close()
        except Exception as e:
            st.error("Error loading Coral Gables data: {}".format(e))

    return firms, brokers, contacts


def render_brokers_page():
    import pandas as pd
    from utils.style import inject_css, LOGO_B64
    inject_css()

    if st.sidebar.button("Back to Home", key="brk_back"):
        st.session_state["app_mode"] = "home"
        st.rerun()

    logo_tag = ""
    if LOGO_B64:
        logo_tag = '<img src="data:image/png;base64,{}" style="height:50px;">'.format(LOGO_B64)
    st.markdown(
        '<div style="display:flex;align-items:center;gap:1rem;margin-bottom:1.5rem;">'
        '{logo}'
        '<div><h2 style="margin:0;color:var(--text-primary);">Broker Database</h2>'
        '<div style="font-size:0.75rem;color:var(--text-muted);">South Florida CRE Contacts &mdash; Brokers, Firms & Lease Prospects</div></div>'
        '</div>'.format(logo=logo_tag),
        unsafe_allow_html=True,
    )

    with st.spinner("Loading contacts..."):
        try:
            firms, brokers, contacts = _load_brokers()
        except Exception as e:
            st.error("Error loading broker data: {}".format(e))
            firms, brokers, contacts = [], [], []

    # KPIs
    st.markdown(
        '<div style="display:flex;gap:1rem;flex-wrap:wrap;margin-bottom:1.5rem;">'
        '<div class="kpi-card"><div class="kl">CRE Firms</div><div class="kv">{}</div></div>'
        '<div class="kpi-card"><div class="kl">Brokers</div><div class="kv">{}</div></div>'
        '<div class="kpi-card"><div class="kl">Lease Contacts</div><div class="kv">{}</div></div>'
        '<div class="kpi-card"><div class="kl">Total Contacts</div><div class="kv">{}</div></div>'
        '</div>'.format(len(firms), len(brokers), len(contacts), len(firms) + len(brokers) + len(contacts)),
        unsafe_allow_html=True,
    )

    tab1, tab2, tab3 = st.tabs(["Brokers ({})".format(len(brokers)), "CRE Firms ({})".format(len(firms)), "Lease Contacts ({})".format(len(contacts))])

    # ── Tab 1: Brokers ──
    with tab1:
        if brokers:
            # Search
            search = st.text_input("Search brokers", placeholder="Name, organization, type...", key="brk_search")

            filtered = brokers
            if search:
                q = search.lower()
                filtered = [b for b in brokers if q in b["name"].lower() or q in b["organization"].lower() or q in b["type"].lower()]

            # Organization filter
            orgs = sorted(set(b["organization"] for b in filtered if b["organization"] and b["organization"] != "None"))
            org_filter = st.selectbox("Filter by organization", ["All"] + orgs, key="brk_org")
            if org_filter != "All":
                filtered = [b for b in filtered if b["organization"] == org_filter]

            st.markdown('<div style="font-size:0.72rem;color:var(--text-muted);margin-bottom:0.5rem;">{} brokers</div>'.format(len(filtered)), unsafe_allow_html=True)

            rows = []
            for b in filtered:
                linkedin = ""
                if b["linkedin"] and b["linkedin"] not in ("/", "/ ", "None", ""):
                    lnk = b["linkedin"].strip()
                    if not lnk.startswith("http"):
                        lnk = "https://" + lnk
                    linkedin = lnk
                rows.append({
                    "Name": b["name"],
                    "Organization": b["organization"],
                    "Type": b["type"],
                    "Email": b["email"],
                    "Mobile": b["mobile"],
                    "Market": b["market"],
                    "LinkedIn": linkedin,
                })
            df = pd.DataFrame(rows)
            st.dataframe(df.astype(str), use_container_width=True, hide_index=True)
        else:
            st.info("No broker data found.")

    # ── Tab 2: CRE Firms ──
    with tab2:
        if firms:
            rows = []
            for f in firms:
                rows.append({
                    "Firm": f["firm"],
                    "Category": f["category"],
                    "Website": f["website"],
                    "Status": f["status"],
                })
            df = pd.DataFrame(rows)
            st.dataframe(df.astype(str), use_container_width=True, hide_index=True)
        else:
            st.info("No firm data found.")

    # ── Tab 3: Lease Contacts ──
    with tab3:
        if contacts:
            search2 = st.text_input("Search contacts", placeholder="Name, type, market...", key="ct_search")
            filtered2 = contacts
            if search2:
                q = search2.lower()
                filtered2 = [c for c in contacts if q in c["name"].lower() or q in c.get("type", "").lower() or q in c.get("market", "").lower() or q in c.get("organization", "").lower()]

            # Market filter
            markets = sorted(set(c.get("market", "") for c in filtered2 if c.get("market") and c["market"] != "None"))
            mkt_filter = st.selectbox("Filter by market/area", ["All"] + markets, key="ct_mkt")
            if mkt_filter != "All":
                filtered2 = [c for c in filtered2 if c.get("market") == mkt_filter]

            st.markdown('<div style="font-size:0.72rem;color:var(--text-muted);margin-bottom:0.5rem;">{} contacts</div>'.format(len(filtered2)), unsafe_allow_html=True)

            rows = []
            for c in filtered2:
                rows.append({
                    "Name": c["name"],
                    "Organization": c.get("organization", ""),
                    "Type": c.get("type", ""),
                    "Email": c.get("email", ""),
                    "Mobile": c.get("mobile", ""),
                    "Market": c.get("market", ""),
                    "Address": c.get("address", ""),
                })
            df = pd.DataFrame(rows)
            st.dataframe(df.astype(str), use_container_width=True, hide_index=True)
        else:
            st.info("No contact data found.")
